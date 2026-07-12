#!/usr/bin/env python3
"""
sync_boletos_calculadora.py
Platea · El Gorila S2

Qué hace:
  Lee /api/{tid}/disponibilidad de la boletera (mismo endpoint que ya usa el
  Agente 12 — Boletera Sync) para cada una de las 11 funciones de S2, y escribe
  "boletos vendidos" + "% ocupación" en las columnas N y O de
  Calculadora_Reparto_El_Gorila.xlsx (columnas auxiliares, no alimentan ninguna
  fórmula de dinero).

Qué NO hace (limitación real, no de este script):
  NO llena "Ingreso taquilla" (columna C). La boletera hoy no expone ingreso
  real ni desglose por tipo de boleto — /api/reporte (worker de solo lectura)
  tiene hardcodeado ingresos_mxn=0 y por_tipo={} porque ese dato vive en Stripe,
  no en el KV de inventario. Ver hoja "Notas y Riesgos" del Excel, punto 1.
  Mientras eso no se resuelva, Dirección sigue escribiendo el ingreso a mano (dato
  exacto que ya ve en su dashboard de Stripe).

Uso:
  python sync_boletos_calculadora.py [--excel /ruta/al/archivo.xlsx]

Env (.env, mismos nombres que Agente 12):
  BOLETERA_URL   → https://elgorila-api.dupeyronosterlen.workers.dev
  TEATRO_ID      → gorila
"""

import os
import sys
import datetime
import requests
from openpyxl import load_workbook

BOLETERA_URL = os.getenv("BOLETERA_URL", "https://elgorila-api.dupeyronosterlen.workers.dev")
TEATRO_ID = os.getenv("TEATRO_ID", "gorila")
DEFAULT_EXCEL = "/Volumes/La Mancha/Elgorila/usuario/datos/Calculadora_Reparto_El_Gorila.xlsx"
SHEET = "Reparto por Función"

FUNCIONES_S2 = [
    "2026-07-18", "2026-07-25", "2026-08-01", "2026-08-08", "2026-08-15",
    "2026-08-22", "2026-08-29", "2026-09-05", "2026-09-12", "2026-09-19", "2026-09-26",
]


def get_disponibilidad(fecha_str: str) -> dict | None:
    try:
        r = requests.get(
            f"{BOLETERA_URL}/api/{TEATRO_ID}/disponibilidad",
            params={"fecha": fecha_str}, timeout=10,
        )
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"   ⚠️  {fecha_str}: no se pudo leer ({e})")
        return None


def main():
    excel_path = DEFAULT_EXCEL
    if "--excel" in sys.argv:
        excel_path = sys.argv[sys.argv.index("--excel") + 1]

    if not os.path.exists(excel_path):
        print(f"❌ No existe el archivo: {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path)
    ws = wb[SHEET]

    # Encontrar la fila de cada fecha buscando en la columna B (fecha función)
    fecha_to_row = {}
    for row in ws.iter_rows(min_row=1, max_col=2):
        val = row[1].value  # columna B
        if isinstance(val, str) and val in FUNCIONES_S2:
            fecha_to_row[val] = row[0].row

    print(f"🎟️  Sync boletos → {excel_path}")
    print(f"   Boletera: {BOLETERA_URL} (tid={TEATRO_ID})\n")

    actualizados = 0
    for fecha in FUNCIONES_S2:
        rr = fecha_to_row.get(fecha)
        if rr is None:
            print(f"   ⚠️  {fecha}: no encontré la fila en el Excel — ¿cambiaste la estructura?")
            continue
        disp = get_disponibilidad(fecha)
        if disp is None:
            continue
        vendidos = int(disp.get("vendidos", 0))
        total = int(disp.get("total", 0)) or 280
        ocupacion = round(vendidos / total, 3) if total else 0
        ws.cell(row=rr, column=14, value=vendidos)       # N
        ws.cell(row=rr, column=15, value=ocupacion)       # O
        print(f"   ✅ {fecha}: {vendidos}/{total} vendidos ({ocupacion*100:.1f}%)")
        actualizados += 1

    wb.save(excel_path)
    print(f"\n💾 Guardado. {actualizados}/{len(FUNCIONES_S2)} funciones actualizadas.")
    print("   Recuerda: 'Ingreso taquilla' (columna C) sigue siendo manual — ver README arriba.")


if __name__ == "__main__":
    main()
